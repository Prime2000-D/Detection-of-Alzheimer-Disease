import cv2
import skimage.feature as sk
import numpy as np
import openpyxl as ox
import xlrd
import ast
from skimage.io import imread_collection

#im= cv2.imread('IM-0115-0001.jpeg')#bring image into python
im_col= imread_collection('C:\\Users\\Prime\\Desktop\\Detection of Alzhemier Disease project\\dataset_final\\MildDemented\\*')
#print(im_col)
print(type(im_col))#stored im_col data to variable

r = 1440

for i in range (0, 717):

    image=im_col[i]
    
    try:  
        gray=cv2.cvtColor(image,cv2.COLOR_BGR2GRAY)#coverted to grayscale
        img=np.array(gray)#turn into array
        greycomat=sk.greycomatrix(img,[1],[0],256,symmetric=True,normed=True) #create glcm
        #print(greycomat)
        #thing = greycomat.tofile("GLCMmatrix.txt",format = "%s",sep = " " )

        vari = np.var(img)
        print(vari)

        meanie = np.mean(img)
        print(meanie)

        moment = cv2.moments(img)
        print(moment)

        a = sk.greycoprops(greycomat, 'contrast')
        b = sk.greycoprops(greycomat, 'ASM')
        c = sk.greycoprops(greycomat, 'dissimilarity')
        d = sk.greycoprops(greycomat, 'homogeneity')
        e = sk.greycoprops(greycomat, 'correlation')
        f = sk.greycoprops(greycomat, 'energy')

        a_arr = np.array(a)
        b_arr = np.array(b)
        c_arr = np.array(c)
        d_arr = np.array(d)
        e_arr = np.array(e)
        f_arr = np.array(f)

        a_sc = np.asscalar(a)
        b_sc = np.asscalar(b)
        c_sc = np.asscalar(c)
        d_sc = np.asscalar(d)
        e_sc = np.asscalar(e)
        f_sc = np.asscalar(f)

        print(a_sc)
        print(b_sc)
        print(c_sc)
        print(d_sc)
        print(e_sc)
        print(f_sc)

        m = (vari, meanie, a_sc, b_sc, c_sc, d_sc, e_sc, f_sc, "Very Mild Demented")
        filename = "C:\\Users\\Prime\\Desktop\\Detection of Alzhemier Disease project\\feature-extraction.xlsx"
        wb= ox.load_workbook(filename)

        print(wb)
        for ws in wb.worksheets:
            print(ws.title)

        ws = wb.worksheets[0]
        print("**************************************************************************************",i+1,"**************************************************************************************")
        for j in range(len(m)):
            ws.cell(row=r, column=j+1).value = m[j]
        r = r+1
        
        #for i in range (1):
         #   for j in range (len(m)):
          #      ws.cell(row =1,column= j+1, value = m[j])
           #     j+1

        wb.save(filename)

    except:
        img=np.array(image)#turn into array
        greycomat=sk.greycomatrix(img,[1],[0],256,symmetric=True,normed=True)#create glcm
        #print(greycomat)
        #thing = greycomat.tofile("GLCMmatrix.txt",format = "%s",sep = " " )

        vari = np.var(img)
        print(vari)

        meanie = np.mean(img)
        print(meanie)

        moment = cv2.moments(img)
        print(moment)

        a = sk.greycoprops(greycomat, 'contrast')
        b = sk.greycoprops(greycomat, 'ASM')
        c = sk.greycoprops(greycomat, 'dissimilarity')
        d = sk.greycoprops(greycomat, 'homogeneity')
        e = sk.greycoprops(greycomat, 'correlation')
        f = sk.greycoprops(greycomat, 'energy')

        a_arr = np.array(a)
        b_arr = np.array(b)
        c_arr = np.array(c)
        d_arr = np.array(d)
        e_arr = np.array(e)
        f_arr = np.array(f)

        a_sc = np.asscalar(a)
        b_sc = np.asscalar(b)
        c_sc = np.asscalar(c)
        d_sc = np.asscalar(d)
        e_sc = np.asscalar(e)
        f_sc = np.asscalar(f)

        print(a_sc)
        print(b_sc)
        print(c_sc)
        print(d_sc)
        print(e_sc)
        print(f_sc)

        m = (vari, meanie, a_sc, b_sc, c_sc, d_sc, e_sc, f_sc, "Very Mild Demented")
        filename = "C:\\Users\\Prime\\Desktop\\Detection of Alzhemier Disease project\\feature-extraction.xlsx"
        wb= ox.load_workbook(filename)

        print(wb)
        for ws in wb.worksheets:
            print(ws.title)

        ws = wb.worksheets[0]
        print("**************************************************************************************",i+1,"**************************************************************************************")
        for j in range(len(m)):
            ws.cell(row=r, column=j+1).value = m[j]
        r = r+1
        
        #for i in range (1):
         #   for j in range (len(m)):
          #      ws.cell(row =1,column= j+1, value = m[j])
           #     j+1

        wb.save(filename)

