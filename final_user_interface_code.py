import tkinter as tk
from tkinter import *
from tkinter import filedialog
from PIL import Image
import cv2
import skimage.feature as sk
import numpy as np
import openpyxl as ox
import xlrd
import ast
from skimage.io import imread_collection
import pandas
import pickle
from PIL import ImageTk,Image
from tkinter import messagebox
#window = tkinter.Tk()#creates a new window
window = tk.Tk()
window.title("Alzheimer detection")

def clicked():
    #global img
    filename = filedialog.askopenfilename()
    #im2=Image.open(filename)#this processes the image as an object
    #print(img)    
    im= cv2.imread(filename)#this processes the image as an array
    print(im)
    try:             

        gray=cv2.cvtColor(im,cv2.COLOR_BGR2GRAY)#coverted to grayscale
        img=np.array(gray)#turn into array
        greycomat=sk.greycomatrix(img,[1],[0],256,symmetric=True,normed=True)#create glcm
        #print(greycomat)
        #thing = greycomat.tofile("GLCMmatrix.txt",format = "%s",sep = " " )
        vari = np.var(img)
        print(vari)
        meanie = np.mean(img)
        print(meanie)
        moment = cv2.moments(img)
        print(moment)
        contrastt = sk.greycoprops(greycomat, 'contrast')
        asmm = sk.greycoprops(greycomat, 'ASM')
        dissimlarityy = sk.greycoprops(greycomat, 'dissimilarity')
        homogeneityy = sk.greycoprops(greycomat, 'homogeneity')
        correlationn = sk.greycoprops(greycomat, 'correlation')
        energyy = sk.greycoprops(greycomat, 'energy')
        a_arr = np.array(contrastt)
        b_arr = np.array(asmm)
        c_arr = np.array(dissimlarityy)
        d_arr = np.array(homogeneityy)
        e_arr = np.array(correlationn)
        f_arr = np.array(energyy)
        a_sc = np.asscalar(contrastt)
        b_sc = np.asscalar(asmm)
        c_sc = np.asscalar(dissimlarityy)
        d_sc = np.asscalar(homogeneityy)
        e_sc = np.asscalar(correlationn)
        f_sc = np.asscalar(energyy)
        print(a_sc)#contrast
        print(b_sc)#asm
        print(c_sc)#dissimilarity
        print(d_sc)#homogeneity
        print(e_sc)#correlation
        print(f_sc)#energy
        m = (vari, meanie, a_sc, b_sc, c_sc, d_sc, e_sc, f_sc)
        filename = "C:\\Users\\Prime\\Desktop\\Detection of Alzhemier Disease project\\test-image.xlsx"
        wb= ox.load_workbook(filename)
        print(wb)
        for ws in wb.worksheets:
            print(ws.title)

        ws = wb.worksheets[0]      
        for j in range(len(m)):
            ws.cell(row=2, column=j+1).value = m[j]
       
        wb.save(filename)
    except:
        im=np.array(im)#turn into array
        greycomat=sk.greycomatrix(im,[1],[0],256,symmetric=True,normed=True)#create glcm
        #print(greycomat)
        #thing = greycomat.tofile("GLCMmatrix.txt",format = "%s",sep = " " )
        vari = np.var(im)
        print(vari)
        meanie = np.mean(im)
        print(meanie)
        moment = cv2.moments(im)
        print(moment)
        a = sk.greycoprops(greycomat, 'contrast')
        b = sk.greycoprops(greycomat, 'ASM')
        c = sk.greycoprops(greycomat, 'dissimilarity')
        d = sk.greycoprops(greycomat, 'homogeneity')
        e = sk.greycoprops(greycomat, 'correlation')
        f = sk.greycoprops(greycomat, 'energy')
        a_arr = np.array(a)
        b_arr = np.array(b)
        c_arr = np.array(c)
        d_arr = np.array(d)
        e_arr = np.array(e)
        f_arr = np.array(f)
        a_sc = np.asscalar(a)
        b_sc = np.asscalar(b)
        c_sc = np.asscalar(c)
        d_sc = np.asscalar(d)
        e_sc = np.asscalar(e)
        f_sc = np.asscalar(f)
        print(a_sc)
        print(b_sc)
        print(c_sc)
        print(d_sc)
        print(e_sc)
        print(f_sc)
        m = (vari, meanie, a_sc, b_sc, c_sc, d_sc, e_sc, f_sc)
        filename = "C://Users//SHASHU//Downloads//test-image.xlsx"
        wb= ox.load_workbook(filename)

        print(wb)
        for ws in wb.worksheets:
            print(ws.title)

        ws = wb.worksheets[0]
        for j in range(len(m)):
            ws.cell(row=1, column=j+1).value = m[j]
        wb.save(filename)
    data = pandas.read_excel("C:\\Users\\Prime\\Desktop\\Detection of Alzhemier Disease project\\test-image.xlsx",engine='openpyxl')
    print("features of the image are")
    print(data)
    svmmodel="C:\\Users\\Prime\\Desktop\\Detection of Alzhemier Disease project\\finalized_model_svm.sav"
    loaded_model_svm = pickle.load(open(svmmodel, 'rb'))
    prediction=loaded_model_svm.predict(data)
    print(prediction)
    #thelabel = Label(window, text = prediction)
    #thelabel.pack()
    messagebox.showinfo(title="Diagnosis", message=prediction)
b = Button(window, text = "Submit", command=clicked,font="Arial" )
b.pack()
#b.grid(row = 8, column = 3)
menubar = Menu(window)
filemenu = Menu(menubar, tearoff=0)
filemenu.add_command(label="New")#way of creating commands for wihtin one option
filemenu.add_command(label="Save Result")
filemenu.add_command(label="Save Result As..")
filemenu.add_separator()

filemenu.add_command(label="Exit", command=window.quit)
menubar.add_cascade(label="File", menu=filemenu)
window.config(menu=menubar)
explanation="Click on the submit button. This will open a file explorer window."+"\n"+" Navigate through the file explorer to select the MRI image that requires diagnosis."+"\n"+ "Upon selecting the image, the diagnosis will be acquired within 10 seconds."
systemdirec=Label(window, text= explanation)
systemdirec.pack()

window.mainloop()
